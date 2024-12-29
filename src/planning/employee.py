"""Employee module."""

import json
import os
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
import yaml
from openpyxl import (
    load_workbook,
)
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils.dataframe import dataframe_to_rows
from pandas.core.indexes.frozen import FrozenList


def create_employees(employee_restrictions):
    """Create employees.

    :param employee_restrictions: Dictionary with employee restrictions
    :return:
    """
    employees_full = [
        {
            "capacity": 1,
        }
    ]

    employees_partial = [
        {
            "capacity": 0.77,
        }
    ]

    employees_temp = employees_full * 2 + employees_partial * 1

    employees = []
    for index, one_employee in enumerate(employees_temp):
        employees.append(
            {
                "name": f"E{index + 1}",
                "capacity": one_employee["capacity"],
                "max_hours_year": employee_restrictions["max_hours_year_employee"] * one_employee["capacity"],
                "max_hours_week": employee_restrictions["max_hours_week_employee"] * one_employee["capacity"],
            }
        )

    return employees


def create_employees_with_dates(start_date, num_days, employees):
    """Create employees with dates.

    :param start_date: First date of the year
    :param num_days: Number of days to generate
    :param employees: List of employees
    :return:
    """
    dates = pd.date_range(start=start_date, periods=num_days, freq="D")
    employees_info = pd.DataFrame(index=dates, columns=[emp for emp in employees.keys()], data="")
    return employees_info, dates


def init_employees_by_shifts(dates, employee_restrictions):
    """Init employees by shifts.

    :param dates: List of dates
    :param employee_restrictions: Dictionary with employee restrictions
    :return:
    """
    all_employees_by_shift = pd.DataFrame(
        index=dates,
        columns=[one_shift for one_shift in employee_restrictions["shifts"]],
    )
    all_employees_by_shift[:] = 0

    return all_employees_by_shift


def get_weekends_of_month(year, month):
    """Get weekends of month.

    :param year: Year as an integer
    :param month: Month as an integer
    :return: DataFrame with weekends of the month
    """
    start_date = pd.Timestamp(year=year, month=month, day=1)
    end_date = start_date + pd.offsets.MonthEnd(1)
    date_range = pd.date_range(start=start_date, end=end_date, freq="D")

    weekends = date_range[date_range.weekday.isin([4, 5, 6])]

    weekends_df = pd.DataFrame(weekends, columns=["Date"])

    num_weekends = len(weekends_df) // 2

    return weekends_df, num_weekends


def count_weekend_workdays(employees_info, employee, year, month):
    """Count the number of weekend workdays for the given employee in the specified month.

    :param employees_info: DataFrame with employee information
    :param employee: Employee name or ID
    :param year: Year as an integer
    :param month: Month as an integer
    :return: Number of weekend workdays
    """
    weekends_df = get_weekends_of_month(year, month)

    weekends_df, _ = weekends_df
    weekend_workdays = employees_info.loc[weekends_df["Date"], employee]
    total_weekend_workdays = weekend_workdays.value_counts().reindex(["M", "T"], fill_value=0).sum()

    return total_weekend_workdays


def count_remaining_weekends(date):
    """Count the number of remaining weekends in the current month from the given date.

    :param date: Date as a pandas Timestamp or a string in 'YYYY-MM-DD' format
    :return: Number of remaining weekends in the current month
    """
    if isinstance(date, str):
        date = pd.Timestamp(date)

    weekends_df, _ = get_weekends_of_month(date.year, date.month)

    remaining_weekends = weekends_df[weekends_df["Date"] > date]

    if isinstance(weekends_df, pd.DataFrame) and "Date" in weekends_df.columns:
        remaining_weekends = weekends_df[weekends_df["Date"] > date]

        num_remaining_weekends = len(remaining_weekends) // 2
    else:
        num_remaining_weekends = 0

    return num_remaining_weekends


def get_current_week_dates(date, start_date):
    """Get all dates of the current week for a given date.

    :param date: Date as a pandas Timestamp or a string in 'YYYY-MM-DD' format
    :param start_date: First date of the year
    :return: List of dates in the current week
    """
    if isinstance(date, str):
        date = pd.Timestamp(date)

    start_of_week = date - pd.Timedelta(days=date.weekday())
    week_dates = pd.date_range(start=start_of_week, end=date, freq="D")

    if start_of_week < pd.Timestamp(start_date):
        start_of_week = pd.Timestamp(start_date)
        week_dates = pd.date_range(start=start_of_week, end=date, freq="D")

    return week_dates


def count_week_restdays(employees_info, employee, date, start_date):
    """Count the number of rest days ("-" and "V") for the given employee in the current week.

    :param employees_info: DataFrame with employee information
    :param employee: Employee name or ID
    :param date: Date as a pandas Timestamp or a string in 'YYYY-MM-DD' format
    :param start_date: First date of the year
    :return: Number of rest days in the current week
    """
    week_dates = get_current_week_dates(date, start_date)
    week_restdays = employees_info.loc[week_dates, employee]
    total_week_restdays = week_restdays.value_counts().reindex(["-", "V"], fill_value=0).sum()

    return total_week_restdays


def count_weekend_restdays(employees_info, employee, year, month):
    """Count the number of weekend restdays for the given employee in the specified month.

    :param employees_info: DataFrame with employee information
    :param employee: Employee name or ID
    :param year: Year as an integer
    :param month: Month as an integer
    :return: Number of weekend workdays
    """
    weekends_df, _ = get_weekends_of_month(year, month)

    weekend_workdays = employees_info.loc[weekends_df["Date"], employee]
    total_weekend_workdays = weekend_workdays.value_counts().reindex(["-", "V"], fill_value=0).sum()

    return total_weekend_workdays


def get_previous_day_value(employees_info, employee, date):
    """Get the value for the previous day in employees_info for a given date and employee.

    :param employees_info: DataFrame with employee information
    :param employee: Employee name or ID
    :param date: Date as a pandas Timestamp or a string in 'YYYY-MM-DD' format
    :return: Value for the previous day
    """
    if isinstance(date, str):
        date = pd.Timestamp(date)

    previous_day = date - pd.Timedelta(days=1)

    if previous_day in employees_info.index:
        previous_day_value = employees_info.loc[previous_day, employee]
    else:
        previous_day_value = None

    return previous_day_value


def skip_employee(one_employee, available_employees, date, num_remaining_weekends, any_employee_rest_in_weekend):
    """Evaluate if an employee should be skipped based on various conditions.

    :param one_employee: Dictionary with employee information
    :param available_employees: List of available employees
    :param date: Date as a pandas Timestamp or a string in 'YYYY-MM-DD' format
    :param num_remaining_weekends: Number of remaining weekends in the current month
    :return: True if the employee should be skipped, False otherwise
    """
    if isinstance(date, str):
        date = pd.Timestamp(date)

    return len(available_employees) > 1 and is_weekend(date) and not any_employee_rest_in_weekend


def assign_employee_shift(date, shift, one_employee, all_employees_by_shift, employees_info):
    """Assign a shift to an employee and update the DataFrames.

    :param date: Date as a pandas Timestamp or a string in 'YYYY-MM-DD' format
    :param shift: Shift to be assigned
    :param one_employee: Dictionary with employee information
    :param all_employees_by_shift: DataFrame tracking the number of employees by shift
    :param employees_info: DataFrame with employee information
    """
    if isinstance(date, str):
        date = pd.Timestamp(date)

    all_employees_by_shift.loc[date, shift] += 1
    employees_info.loc[date, one_employee["employee"]] = shift


def sort_available_employees(available_employees, shift, month, data_employee_monthly):
    """Sort available employees based on specific criteria.

    :param available_employees: List of available employees
    :param shift: Shift to be assigned
    :return: Sorted list of available employees
    """
    return sorted(
        available_employees,
        key=lambda x: (
            x["previous_day_value"] != shift,
            data_employee_monthly[month][x["employee"]]["rest_weekends"],
        ),
        reverse=False,
    )


def assign_available_employees(
    date,
    shift,
    available_employees,
    all_employees_by_shift,
    employees_info,
    employee_restrictions,
    num_remaining_weekends,
    any_employee_rest_in_weekend,
    data_employee_monthly,
):
    """Assign shifts to available employees.

    :param date: Date as a pandas Timestamp or a string in 'YYYY-MM-DD' format
    :param shift: Shift to be assigned
    :param available_employees: List of available employees
    :param all_employees_by_shift: DataFrame tracking the number of employees by shift
    :param employees_info: DataFrame with employee information
    :param employee_restrictions: Dictionary with employee restrictions
    :param num_remaining_weekends: Number of remaining weekends in the current month
    """
    for one_employee in available_employees:
        if skip_employee(
            one_employee,
            available_employees,
            date,
            num_remaining_weekends,
            any_employee_rest_in_weekend.get(shift, False),
        ):
            available_employees = [emp for emp in available_employees if emp["employee"] != one_employee["employee"]]
            if date.weekday() in (4,):
                saturday = date + timedelta(days=1)
                sunday = date + timedelta(days=2)
                employees_info.loc[saturday, one_employee["employee"]] = "-"
                employees_info.loc[sunday, one_employee["employee"]] = "-"
                data_employee_monthly[date.month][one_employee["employee"]]["rest_weekends"] += 1
                any_employee_rest_in_weekend[shift] = True
            continue
        assign_employee_shift(date, shift, one_employee, all_employees_by_shift, employees_info)

        if all_employees_by_shift.loc[date, shift] >= employee_restrictions["max_persons_per_shift"][shift]:
            break  # No more employees needed


def is_weekend(date):
    """Check if a given date is a weekend.

    :param date: Date as a pandas Timestamp or a string in 'YYYY-MM-DD' format
    :return: True if the date is a weekend (Saturday or Sunday), False otherwise
    """
    return date.weekday() in (4, 5, 6)


def load_data_by_date(all_employees_by_shift, employee_restrictions, employees_info, employees, start_date):
    """Load data by date.

    :param all_employees_by_shift: DataFrame tracking the number of employees by shift
    :param employee_restrictions: Dictionary with employee restrictions
    :param employees_info: DataFrame with employee information
    :param employees: List of employees
    :param start_date: First date of the year
    :return:
    """
    data_employee_monthly = {}
    any_employee_rest_in_weekend = {}
    for date in all_employees_by_shift.index:
        num_remaining_weekends = count_remaining_weekends(date)
        for shift in all_employees_by_shift.columns:
            if (
                all_employees_by_shift.loc[date, shift] >= employee_restrictions["max_persons_per_shift"][shift]
            ):  # No more employees needed
                continue

            available_employees = []
            if date.weekday() in (4,):
                any_employee_rest_in_weekend[shift] = False

            for employee in employees_info.columns:
                month = date.month
                if month not in data_employee_monthly:
                    data_employee_monthly[month] = {}
                if employee not in data_employee_monthly[month]:
                    data_employee_monthly[month][employee] = {
                        "rest_weekends": 0,
                    }

                if not employees_info.loc[date, employee] or employees_info.loc[date, employee] == "":
                    six_days_ago = date - timedelta(days=6)
                    yesterday = date - timedelta(days=1)

                    last_6_days_employee = employees_info.loc[six_days_ago:date, employee]
                    previous_day_value = get_previous_day_value(employees_info, employee, date)

                    if yesterday in employees_info.index:
                        value_yesterday = employees_info.loc[yesterday, employee]
                    else:
                        value_yesterday = None

                    total_worked_days_in_6_days = (
                        last_6_days_employee.value_counts().reindex(["M", "T"], fill_value=0).sum()
                    )

                    total_sum_m_t = employees_info[employee].isin(["M", "T"]).sum()

                    employee_capacity = next(
                        emp["capacity"] for key_emp, emp in employees.items() if key_emp == employee
                    )
                    if value_yesterday and value_yesterday == "T" and shift == "M":
                        employees_info.loc[date, employee] = ""
                    elif (
                        (total_sum_m_t * employee_restrictions["hours_per_shift"])
                        >= (employee_restrictions["max_hours_year_employee"] * employee_capacity)
                    ) or (
                        ((total_worked_days_in_6_days) * employee_restrictions["hours_per_shift"])
                        >= (employee_restrictions["max_hours_week_employee"])
                    ):
                        employees_info.loc[date, employee] = ""
                    else:
                        available_employees.append(
                            {
                                "employee": employee,
                                "previous_day_value": previous_day_value,
                            }
                        )

            available_employees = sort_available_employees(available_employees, shift, month, data_employee_monthly)

            assign_available_employees(
                date,
                shift,
                available_employees,
                all_employees_by_shift,
                employees_info,
                employee_restrictions,
                num_remaining_weekends,
                any_employee_rest_in_weekend,
                data_employee_monthly,
            )

            num_employees_in_shift = employees_info.loc[date].value_counts().get(shift, 0)
            if num_employees_in_shift < employee_restrictions["min_persons_per_shift"][shift]:
                for one_employee in available_employees:
                    assign_employee_shift(date, shift, one_employee, all_employees_by_shift, employees_info)
                for employee_key in employees.keys():
                    num_worked_days_in_shift = (
                        employees_info.loc[six_days_ago:date, employee_key].value_counts().get(shift, 0)
                    )
                    if (
                        num_worked_days_in_shift > 0
                        and (num_worked_days_in_shift + 1) * employee_restrictions["hours_per_shift"]
                        < employee_restrictions["max_hours_week_employee"]
                        and data_employee_monthly[month][employee_key]["rest_weekends"]
                        and (employees_info.loc[yesterday, employee_key] == "-" and date.weekday() not in (5, 6))
                    ):
                        all_employees_by_shift.loc[date, shift] += 1
                        employees_info.loc[date, employee_key] = shift
                    if (
                        employees_info.loc[date].value_counts().get(shift, 0)
                        >= employee_restrictions["min_persons_per_shift"][shift]
                    ):
                        break
        for one_employee in employees_info.columns:
            if employees_info.loc[date, one_employee] == "":
                employees_info.loc[date, one_employee] = "-"

    return employees_info


def modify_index_to_datetime(dataframe):
    """Modify dataframe index to datetime.

    :param dataframe:
    :return:
    """
    dataframe.index = pd.to_datetime(dataframe.index)
    dataframe.index = dataframe.index.strftime("%Y-%m-%d")


def generate_excel(dataframe, filename):
    """Generate excel.

    :param dataframe:
    :param filename:
    :return:
    """
    output_filename = filename
    dataframe.to_excel(output_filename, sheet_name="Shift Schedule")


def load_translations():
    """Load translations.

    :return:
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    lang_file_path = os.path.join(script_dir, "lang.yaml")
    with open(lang_file_path) as file:
        lang_data = yaml.safe_load(file)
    return lang_data


def create_transposed_dataframe(employees_info, lang="es"):
    """Create a transposed dataframe.

    :param employees_info:
    :param lang:
    :return:
    """
    employees_info.index = pd.to_datetime(employees_info.index)

    lang_data = load_translations()

    day_of_month = employees_info.index.day
    days_of_week_map = {i: lang_data["days_of_week"][i][lang] for i in range(7)}
    months_map = {month: lang_data["months"][month][lang] for month in range(1, 13)}
    month = employees_info.index.month.map(months_map)
    day_of_week = employees_info.index.dayofweek.map(days_of_week_map)

    multi_index_index = pd.MultiIndex.from_arrays([month, day_of_week, day_of_month], names=["", "", ""])

    employees_info.index = multi_index_index

    transposed_employees_info = employees_info.T

    return transposed_employees_info


def generate_summary(employees, employee_restrictions, transposed_employees_info):
    """Generate summary.

    :param employees:
    :param employee_restrictions:
    :param transposed_employees_info:
    :return:
    """
    transposed_employees_info["THT"] = transposed_employees_info.apply(
        lambda row: (row.value_counts().get("M", 0) + row.value_counts().get("T", 0) + row.value_counts().get("N", 0))
        * employee_restrictions["hours_per_shift"],
        axis=1,
    )

    transposed_employees_info["MH"] = transposed_employees_info.index.map(lambda emp: employees[emp]["max_hours_year"])
    transposed_employees_info["Diff"] = transposed_employees_info["MH"] - transposed_employees_info["THT"]

    sum_m_t = transposed_employees_info.apply(lambda col: col.isin(["M", "T"]).sum(), axis=0)
    new_row = pd.Series(sum_m_t, name="Total")
    transposed_employees_info = pd.concat([transposed_employees_info, new_row.to_frame().T])

    transposed_employees_info.loc["Total", ["THT", "MH", "Diff"]] = [np.nan] * 3

    return transposed_employees_info


def generate_summary_month(employees, employee_restrictions, planning_data):
    """Generate summary month.

    :param employees:
    :param employee_restrictions:
    :param planning_data:
    :return:
    """
    planning_data["THT"] = planning_data.iloc[1:].apply(
        lambda row: (row.value_counts().get("M", 0) + row.value_counts().get("T", 0) + row.value_counts().get("N", 0))
        * employee_restrictions["hours_per_shift"],
        axis=1,
    )

    sum_m_t = planning_data.apply(lambda col: col.isin(["M", "T", "N"]).sum(), axis=0)
    new_row = pd.Series(sum_m_t, name="Total")
    planning_data = pd.concat([planning_data, new_row.to_frame().T])


def generate_summary_total(employees, employee_restrictions, planning_data):
    total_data = {
        "THT": 0,
        "MHA": 0,
        "Diff": 0,
    }

    for month in range(1, 13):
        month_str = f"{month:02d}"
        generate_summary_month(employees, employee_restrictions, planning_data[month_str])
        total_data["THT"] += planning_data[month_str]["THT"]

    total_data["MHA"] = planning_data["01"].index.map(
        lambda emp: employees[emp]["max_hours_year"] if emp in employees else np.nan
    )

    total_data["Diff"] = total_data["MHA"] - total_data["THT"]

    total_data_df = pd.DataFrame(total_data)

    total_data = total_data_df[total_data_df.index.notnull() & (total_data_df.index != "")]

    return total_data


def generate_transposed_excel_with_styles(transposed_employees_info, employee_restrictions, filename):
    """Generate transposed excel with styles.

    :param transposed_employees_info:
    :param employee_restrictions:
    :param filename:
    :return:
    """
    output_filename = filename
    transposed_employees_info.to_excel(output_filename, sheet_name="Shift Schedule")

    workbook = load_workbook(output_filename)
    worksheet = workbook["Shift Schedule"]

    worksheet.delete_rows(4)

    min_width = 3
    for col in worksheet.iter_cols():
        for cell in col:
            if not any(cell.coordinate in merged_cell for merged_cell in worksheet.merged_cells.ranges):
                column = cell.column_letter
                worksheet.column_dimensions[column].width = min_width
                break

    for i in range(0, 3):
        column_letter = worksheet.cell(row=1, column=worksheet.max_column - i).column_letter
        worksheet.column_dimensions[column_letter].width = 7
        for cell in worksheet[column_letter]:
            cell.alignment = Alignment(horizontal="center")

    first_column_letter = worksheet.cell(row=1, column=1).column_letter
    worksheet.column_dimensions[first_column_letter].width = 7

    fill = PatternFill(start_color="0099FF", end_color="0099FF", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)

    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = font

    weekend_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col in worksheet.iter_cols(min_row=2, max_row=worksheet.max_row, min_col=2, max_col=worksheet.max_column):
        day_of_week_cell = col[0]
        if day_of_week_cell.value in ["S", "D"]:
            for cell in col:
                cell.fill = weekend_fill

    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    min_persons_day = (
        employee_restrictions["min_persons_per_shift"]["M"] + employee_restrictions["min_persons_per_shift"]["T"]
    )

    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    total_row = worksheet.max_row
    for cell in worksheet[total_row]:
        if str(cell.value).isdigit() and int(cell.value) < min_persons_day:
            cell.fill = red_fill

    workbook.save(output_filename)


def assign_vacations(employees_info, vacations_file):
    """Assign vacations to employees.

    :param employees_info: DataFrame with employee information
    :param vacations_file: Path to the vacations file
    """

    with open(vacations_file) as file:
        vacations = yaml.safe_load(file)

    for employee, days in vacations.items():
        if days:
            for day in days:
                day = pd.Timestamp(day)
                if day in employees_info.index:
                    employees_info.loc[day, employee] = "V"


def load_planning_from_yaml(employees_info, planning_file):
    """Load planning from a YAML file.

    :param employees_info:
    :param planning_file:
    """
    with open(planning_file) as file:
        shifts = yaml.safe_load(file)

    for day, shifts_info in shifts.items():
        day = pd.Timestamp(day)
        for shift, employees in shifts_info.items():
            if employees is not None:
                for employee in employees:
                    employees_info.loc[day, employee] = shift


def load_planning_from_xlsx(employees_info, planning_file):
    """Load planning from an Excel file.

    :param employees_info:
    :param planning_file:
    """

    pass


def load_employees_from_yaml(employees_file, employee_restrictions):
    """Load employees from a YAML file.

    :param employees_file:
    :param employee_restrictions:
    :return:
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    working_days_file = os.path.join(script_dir, employees_file)
    with open(working_days_file) as file:
        employees_data = yaml.safe_load(file)

    employees = {}
    for one_employee, one_employee_info in employees_data.items():
        employees.setdefault(one_employee, {})["capacity"] = one_employee_info["capacity"]
        employees[one_employee]["name"] = one_employee_info["name"]
        employees[one_employee]["max_hours_year"] = (
            employee_restrictions["max_hours_year_employee"] * one_employee_info["capacity"]
        )
        employees[one_employee]["max_hours_week"] = (
            employee_restrictions["max_hours_week_employee"] * one_employee_info["capacity"]
        )

    return employees


def export_month(workbook, month_number, planning_data):
    """Export month.

    :param workbook:
    :param month_number:
    :param planning_data:
    """
    month = str(datetime.strptime(month_number, "%m").strftime("%B")).title()

    df = planning_data[month_number]
    df.columns = [datetime.strptime(col, "%d/%m/%y").strftime("%d") if "/" in col else col for col in df.columns]

    worksheet = workbook.create_sheet(title=month)

    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns) + 1)
    worksheet.cell(row=1, column=1, value=month)

    r_idx = 2
    for row in dataframe_to_rows(df, index=True, header=True):
        if isinstance(row, FrozenList):
            continue
        c_idx = 1
        for value in row:
            cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            cell.border = thin_border
            c_idx += 1
        r_idx += 1

    for cell in worksheet[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)
        cell.border = thin_border

    min_width = 3
    for col in worksheet.iter_cols():
        for cell in col:
            if not any(cell.coordinate in merged_cell for merged_cell in worksheet.merged_cells.ranges):
                column = cell.column_letter
                worksheet.column_dimensions[column].width = min_width
                break

    column_letter = worksheet.cell(row=2, column=worksheet.max_column).column_letter
    worksheet.column_dimensions[column_letter].width = 7
    for cell in worksheet[column_letter]:
        cell.alignment = Alignment(horizontal="center")

    first_column_letter = worksheet.cell(row=1, column=1).column_letter
    worksheet.column_dimensions[first_column_letter].width = 3

    fill = PatternFill(start_color="0099FF", end_color="0099FF", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)

    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = font

    weekend_fill = PatternFill(start_color="9CCCE8", end_color="9CCCE8", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col in worksheet.iter_cols(min_row=3, max_row=worksheet.max_row, min_col=2, max_col=worksheet.max_column):
        day_of_week_cell = col[0]
        if day_of_week_cell.value in ["S", "D"]:
            for cell in col:
                cell.fill = weekend_fill

    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")


def add_total_data(workbook, total_data):
    """Add total data.

    :param workbook:
    :param total_data:
    """
    worksheet = workbook.create_sheet(title="Total")

    for r_idx, row in enumerate(dataframe_to_rows(total_data, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
            )
            cell.border = thin_border

    fill = PatternFill(start_color="0099FF", end_color="0099FF", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)

    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = font


def load_config(config_file):
    """Load config from a JSON file.

    :param config_file:
    """
    with open(config_file) as file:
        config = json.load(file)

    return config
